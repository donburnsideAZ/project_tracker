"""
Admin Screen - Manage shared team data and app configuration
Supports dynamic project fields that users can customize
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QScrollArea, QListWidget, QListWidgetItem, QLineEdit,
    QStackedWidget, QFormLayout, QComboBox, QMessageBox,
    QFileDialog, QSplitter, QInputDialog, QDialog, QDialogButtonBox,
    QTableWidget, QTableWidgetItem, QHeaderView, QRadioButton,
    QButtonGroup, QSpinBox, QCheckBox
)
from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QFont
from pathlib import Path
import json
import csv

from ..data_service import data_service


class EditableListPanel(QFrame):
    """A panel for editing a simple list of strings (work types, tags, etc.)"""
    
    data_changed = pyqtSignal()
    
    def __init__(self, title: str, data_path: str):
        super().__init__()
        
        self.title = title
        self.data_path = data_path  # e.g., "work_types", "tags"
        
        self._setup_ui()
        self.refresh()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)
        
        # Header
        header = QHBoxLayout()
        
        title_label = QLabel(self.title)
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title_label.setFont(title_font)
        header.addWidget(title_label)
        
        header.addStretch()
        layout.addLayout(header)
        
        # List widget
        self.list_widget = QListWidget()
        self.list_widget.setAlternatingRowColors(True)
        self.list_widget.itemDoubleClicked.connect(self._on_edit_item)
        layout.addWidget(self.list_widget, 1)
        
        # Add new item row
        add_row = QHBoxLayout()
        
        self.new_item_input = QLineEdit()
        self.new_item_input.setPlaceholderText(f"Add new item...")
        self.new_item_input.returnPressed.connect(self._on_add_item)
        add_row.addWidget(self.new_item_input)
        
        add_btn = QPushButton("+")
        add_btn.setFixedWidth(40)
        add_btn.clicked.connect(self._on_add_item)
        add_row.addWidget(add_btn)
        
        layout.addLayout(add_row)
        
        # Action buttons
        btn_row = QHBoxLayout()
        
        edit_btn = QPushButton("Edit")
        edit_btn.clicked.connect(self._on_edit_selected)
        btn_row.addWidget(edit_btn)
        
        remove_btn = QPushButton("Remove")
        remove_btn.clicked.connect(self._on_remove_item)
        btn_row.addWidget(remove_btn)
        
        import_btn = QPushButton("Import...")
        import_btn.clicked.connect(self._on_import)
        btn_row.addWidget(import_btn)
        
        btn_row.addStretch()
        
        move_up_btn = QPushButton("Up")
        move_up_btn.setFixedWidth(50)
        move_up_btn.clicked.connect(self._on_move_up)
        btn_row.addWidget(move_up_btn)
        
        move_down_btn = QPushButton("Down")
        move_down_btn.setFixedWidth(50)
        move_down_btn.clicked.connect(self._on_move_down)
        btn_row.addWidget(move_down_btn)
        
        layout.addLayout(btn_row)
        
        # Data location hint
        hint = QLabel(f"Data: team_data.json -> {self.data_path}[]")
        hint.setStyleSheet("color: gray; font-size: 10px;")
        layout.addWidget(hint)
    
    def refresh(self):
        """Refresh the list from data service"""
        self.list_widget.clear()
        
        if not data_service.team_data:
            return
        
        items = getattr(data_service.team_data, self.data_path, [])
        for item in items:
            self.list_widget.addItem(item)
    
    def _get_items(self) -> list:
        """Get all items from the list widget"""
        items = []
        for i in range(self.list_widget.count()):
            items.append(self.list_widget.item(i).text())
        return items
    
    def _save(self):
        """Save changes back to team_data"""
        if not data_service.team_data:
            return
        
        setattr(data_service.team_data, self.data_path, self._get_items())
        data_service.save_team_data(data_service.team_data)
        self.data_changed.emit()
    
    def _on_add_item(self):
        """Add a new item"""
        text = self.new_item_input.text().strip()
        if not text:
            return
        
        # Check for duplicates
        existing = self._get_items()
        if text in existing:
            QMessageBox.warning(self, "Duplicate", f"'{text}' already exists.")
            return
        
        self.list_widget.addItem(text)
        self.new_item_input.clear()
        self._save()
    
    def _on_edit_item(self, item: QListWidgetItem):
        """Edit an item on double-click"""
        old_text = item.text()
        new_text, ok = QInputDialog.getText(
            self, 
            "Edit Item", 
            "Value:",
            QLineEdit.EchoMode.Normal,
            old_text
        )
        
        if ok and new_text.strip():
            item.setText(new_text.strip())
            self._save()
    
    def _on_edit_selected(self):
        """Edit the selected item"""
        item = self.list_widget.currentItem()
        if item:
            self._on_edit_item(item)
    
    def _on_remove_item(self):
        """Remove the selected item"""
        row = self.list_widget.currentRow()
        if row < 0:
            return
        
        item = self.list_widget.item(row)
        reply = QMessageBox.question(
            self,
            "Confirm Remove",
            f"Remove '{item.text()}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.list_widget.takeItem(row)
            self._save()
    
    def _on_move_up(self):
        """Move selected item up"""
        row = self.list_widget.currentRow()
        if row > 0:
            item = self.list_widget.takeItem(row)
            self.list_widget.insertItem(row - 1, item)
            self.list_widget.setCurrentRow(row - 1)
            self._save()
    
    def _on_move_down(self):
        """Move selected item down"""
        row = self.list_widget.currentRow()
        if row < self.list_widget.count() - 1:
            item = self.list_widget.takeItem(row)
            self.list_widget.insertItem(row + 1, item)
            self.list_widget.setCurrentRow(row + 1)
            self._save()
    
    def _on_import(self):
        """Import items from CSV, JSON, or Excel file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            f"Import {self.title}",
            "",
            "All Supported (*.json *.csv *.xlsx *.xls);;JSON (*.json);;CSV (*.csv);;Excel (*.xlsx *.xls)"
        )
        
        if not file_path:
            return
        
        try:
            new_items = self._read_import_file(file_path)
            
            if not new_items:
                QMessageBox.information(self, "Import", "No items found in file.")
                return
            
            # Filter out duplicates and existing items
            existing = set(self._get_items())
            added = 0
            skipped = 0
            
            for item in new_items:
                if item and item not in existing:
                    self.list_widget.addItem(item)
                    existing.add(item)
                    added += 1
                else:
                    skipped += 1
            
            if added > 0:
                self._save()
            
            QMessageBox.information(
                self,
                "Import Complete",
                f"Added {added} new items.\nSkipped {skipped} duplicates."
            )
            
        except Exception as e:
            QMessageBox.critical(
                self,
                "Import Error",
                f"Failed to import file:\n{str(e)}"
            )
    
    def _read_import_file(self, file_path: str) -> list:
        """Read items from import file, handling different formats"""
        new_items = []
        
        if file_path.lower().endswith('.json'):
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            if isinstance(data, list):
                for item in data:
                    if isinstance(item, str):
                        new_items.append(item.strip())
                    elif isinstance(item, dict):
                        for key in ['name', 'value', 'label', 'title']:
                            if key in item:
                                new_items.append(str(item[key]).strip())
                                break
        
        elif file_path.lower().endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8', newline='') as f:
                reader = csv.reader(f)
                for row in reader:
                    if row and row[0].strip():
                        new_items.append(row[0].strip())
        
        elif file_path.lower().endswith(('.xlsx', '.xls')):
            new_items = self._import_from_excel(file_path)
        
        return new_items
    
    def _import_from_excel(self, file_path: str) -> list:
        """Import from Excel with column selection dialog"""
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(
                self,
                "Missing Library",
                "Excel import requires openpyxl.\nInstall with: pip install openpyxl"
            )
            return []
        
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active
        
        # Get first 10 rows for preview
        preview_rows = []
        for i, row in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
            preview_rows.append(row)
        
        if not preview_rows:
            return []
        
        # Show column picker dialog
        dialog = ExcelColumnPickerDialog(preview_rows, self)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return []
        
        col_index = dialog.selected_column
        skip_header = dialog.skip_header
        
        # Read selected column
        items = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if skip_header and i == 0:
                continue
            if col_index < len(row) and row[col_index]:
                items.append(str(row[col_index]).strip())
        
        wb.close()
        return items


class ExcelColumnPickerDialog(QDialog):
    """Dialog to select which column to import from Excel"""
    
    def __init__(self, preview_rows: list, parent=None):
        super().__init__(parent)
        self.preview_rows = preview_rows
        self.selected_column = 0
        self.skip_header = True
        
        self.setWindowTitle("Select Column to Import")
        self.setMinimumWidth(500)
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Instructions
        label = QLabel("Select the column containing the values to import:")
        layout.addWidget(label)
        
        # Preview table
        self.table = QTableWidget()
        if self.preview_rows:
            num_cols = max(len(row) for row in self.preview_rows)
            self.table.setColumnCount(num_cols)
            self.table.setRowCount(len(self.preview_rows))
            
            for i, row in enumerate(self.preview_rows):
                for j, val in enumerate(row):
                    item = QTableWidgetItem(str(val) if val else "")
                    self.table.setItem(i, j, item)
            
            self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectColumns)
        self.table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.table.selectColumn(0)
        layout.addWidget(self.table)
        
        # Skip header checkbox
        self.skip_header_check = QPushButton("First row is header (skip it)")
        self.skip_header_check.setCheckable(True)
        self.skip_header_check.setChecked(True)
        layout.addWidget(self.skip_header_check)
        
        # Buttons
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._on_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def _on_accept(self):
        selected = self.table.selectedIndexes()
        if selected:
            self.selected_column = selected[0].column()
        self.skip_header = self.skip_header_check.isChecked()
        self.accept()


class EmployeesPanel(QFrame):
    """Panel for editing employees (has id, name, role)"""
    
    data_changed = pyqtSignal()
    
    def __init__(self):
        super().__init__()
        self._setup_ui()
        self.refresh()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)
        
        # Header
        title_label = QLabel("Employees")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title_label.setFont(title_font)
        layout.addWidget(title_label)
        
        # List widget
        self.list_widget = QListWidget()
        self.list_widget.setAlternatingRowColors(True)
        self.list_widget.itemDoubleClicked.connect(self._on_edit_employee)
        layout.addWidget(self.list_widget, 1)
        
        # Add new employee form
        form_frame = QFrame()
        form_frame.setStyleSheet("""
            QFrame {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 8px;
            }
        """)
        form_layout = QFormLayout(form_frame)
        form_layout.setContentsMargins(8, 8, 8, 8)
        
        self.id_input = QLineEdit()
        self.id_input.setPlaceholderText("OS username (e.g., dsmith)")
        form_layout.addRow("Username:", self.id_input)
        
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Display name")
        form_layout.addRow("Name:", self.name_input)
        
        self.role_combo = QComboBox()
        self.role_combo.setEditable(True)
        self._populate_roles()
        form_layout.addRow("Role:", self.role_combo)
        
        add_btn = QPushButton("+ Add Employee")
        add_btn.clicked.connect(self._on_add_employee)
        form_layout.addRow(add_btn)
        
        layout.addWidget(form_frame)
        
        # Action buttons
        btn_row = QHBoxLayout()
        
        edit_btn = QPushButton("Edit")
        edit_btn.clicked.connect(self._on_edit_selected)
        btn_row.addWidget(edit_btn)
        
        remove_btn = QPushButton("Remove")
        remove_btn.clicked.connect(self._on_remove_employee)
        btn_row.addWidget(remove_btn)
        
        btn_row.addStretch()
        layout.addLayout(btn_row)
        
        # Data location hint
        hint = QLabel("Data: team_data.json -> employees[]")
        hint.setStyleSheet("color: gray; font-size: 10px;")
        layout.addWidget(hint)
    
    def _populate_roles(self):
        """Populate role dropdown from team_roles"""
        self.role_combo.clear()
        if data_service.team_data and data_service.team_data.team_roles:
            self.role_combo.addItems(data_service.team_data.team_roles)
        else:
            self.role_combo.addItems(["Owner", "Contributor", "Reviewer"])
    
    def refresh(self):
        """Refresh the list from data service"""
        self.list_widget.clear()
        self._populate_roles()
        
        if not data_service.team_data:
            return
        
        for emp in data_service.team_data.employees:
            if isinstance(emp, dict):
                display = f"{emp.get('id', '')}  -  {emp.get('name', '')}  [{emp.get('role', '')}]"
                item = QListWidgetItem(display)
                item.setData(Qt.ItemDataRole.UserRole, emp)
                self.list_widget.addItem(item)
    
    def _save(self):
        """Save changes back to team_data"""
        if not data_service.team_data:
            return
        
        employees = []
        for i in range(self.list_widget.count()):
            emp_data = self.list_widget.item(i).data(Qt.ItemDataRole.UserRole)
            employees.append(emp_data)
        
        data_service.team_data.employees = employees
        data_service.save_team_data(data_service.team_data)
        self.data_changed.emit()
    
    def _on_add_employee(self):
        """Add a new employee"""
        emp_id = self.id_input.text().strip()
        name = self.name_input.text().strip()
        role = self.role_combo.currentText().strip()
        
        if not emp_id or not name:
            QMessageBox.warning(self, "Missing Info", "Username and Name are required.")
            return
        
        # Check for duplicate ID
        for i in range(self.list_widget.count()):
            existing = self.list_widget.item(i).data(Qt.ItemDataRole.UserRole)
            if existing.get('id') == emp_id:
                QMessageBox.warning(self, "Duplicate", f"Employee '{emp_id}' already exists.")
                return
        
        emp_data = {'id': emp_id, 'name': name, 'role': role}
        display = f"{emp_id}  -  {name}  [{role}]"
        item = QListWidgetItem(display)
        item.setData(Qt.ItemDataRole.UserRole, emp_data)
        self.list_widget.addItem(item)
        
        # Clear form
        self.id_input.clear()
        self.name_input.clear()
        
        self._save()
    
    def _on_edit_employee(self, item: QListWidgetItem):
        """Edit an employee"""
        emp_data = item.data(Qt.ItemDataRole.UserRole)
        
        dialog = EmployeeEditDialog(emp_data, self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            new_data = dialog.get_data()
            item.setData(Qt.ItemDataRole.UserRole, new_data)
            display = f"{new_data['id']}  -  {new_data['name']}  [{new_data['role']}]"
            item.setText(display)
            self._save()
    
    def _on_edit_selected(self):
        """Edit selected employee"""
        item = self.list_widget.currentItem()
        if item:
            self._on_edit_employee(item)
    
    def _on_remove_employee(self):
        """Remove selected employee"""
        row = self.list_widget.currentRow()
        if row < 0:
            return
        
        item = self.list_widget.item(row)
        emp_data = item.data(Qt.ItemDataRole.UserRole)
        
        reply = QMessageBox.question(
            self,
            "Confirm Remove",
            f"Remove employee '{emp_data.get('name', '')}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.list_widget.takeItem(row)
            self._save()


class EmployeeEditDialog(QDialog):
    """Dialog for editing an employee"""
    
    def __init__(self, emp_data: dict, parent=None):
        super().__init__(parent)
        self.emp_data = emp_data
        
        self.setWindowTitle("Edit Employee")
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QFormLayout(self)
        
        self.id_input = QLineEdit(self.emp_data.get('id', ''))
        layout.addRow("Username:", self.id_input)
        
        self.name_input = QLineEdit(self.emp_data.get('name', ''))
        layout.addRow("Name:", self.name_input)
        
        self.role_combo = QComboBox()
        self.role_combo.setEditable(True)
        if data_service.team_data and data_service.team_data.team_roles:
            self.role_combo.addItems(data_service.team_data.team_roles)
        else:
            self.role_combo.addItems(["Owner", "Contributor", "Reviewer"])
        self.role_combo.setCurrentText(self.emp_data.get('role', ''))
        layout.addRow("Role:", self.role_combo)
        
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
    
    def get_data(self) -> dict:
        return {
            'id': self.id_input.text().strip(),
            'name': self.name_input.text().strip(),
            'role': self.role_combo.currentText().strip()
        }


class TeamRolesPanel(QFrame):
    """Panel for editing team roles that can be assigned to projects"""
    
    data_changed = pyqtSignal()
    
    def __init__(self):
        super().__init__()
        self._setup_ui()
        self.refresh()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)
        
        # Header
        title_label = QLabel("Team Roles")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title_label.setFont(title_font)
        layout.addWidget(title_label)
        
        desc = QLabel("Define roles that can be assigned to team members on projects.")
        desc.setStyleSheet("color: gray;")
        desc.setWordWrap(True)
        layout.addWidget(desc)
        
        # List widget
        self.list_widget = QListWidget()
        self.list_widget.setAlternatingRowColors(True)
        self.list_widget.itemDoubleClicked.connect(self._on_edit_item)
        layout.addWidget(self.list_widget, 1)
        
        # Add new role
        add_row = QHBoxLayout()
        self.new_item_input = QLineEdit()
        self.new_item_input.setPlaceholderText("Add new role...")
        self.new_item_input.returnPressed.connect(self._on_add_item)
        add_row.addWidget(self.new_item_input)
        
        add_btn = QPushButton("+")
        add_btn.setFixedWidth(40)
        add_btn.clicked.connect(self._on_add_item)
        add_row.addWidget(add_btn)
        layout.addLayout(add_row)
        
        # Buttons
        btn_row = QHBoxLayout()
        
        edit_btn = QPushButton("Edit")
        edit_btn.clicked.connect(self._on_edit_selected)
        btn_row.addWidget(edit_btn)
        
        remove_btn = QPushButton("Remove")
        remove_btn.clicked.connect(self._on_remove_item)
        btn_row.addWidget(remove_btn)
        
        btn_row.addStretch()
        layout.addLayout(btn_row)
        
        hint = QLabel("Data: team_data.json -> team_roles[]")
        hint.setStyleSheet("color: gray; font-size: 10px;")
        layout.addWidget(hint)
    
    def refresh(self):
        self.list_widget.clear()
        if not data_service.team_data:
            return
        for role in data_service.team_data.team_roles:
            self.list_widget.addItem(role)
    
    def _get_items(self) -> list:
        return [self.list_widget.item(i).text() for i in range(self.list_widget.count())]
    
    def _save(self):
        if not data_service.team_data:
            return
        data_service.team_data.team_roles = self._get_items()
        data_service.save_team_data(data_service.team_data)
        self.data_changed.emit()
    
    def _on_add_item(self):
        text = self.new_item_input.text().strip()
        if not text:
            return
        if text in self._get_items():
            QMessageBox.warning(self, "Duplicate", f"'{text}' already exists.")
            return
        self.list_widget.addItem(text)
        self.new_item_input.clear()
        self._save()
    
    def _on_edit_item(self, item: QListWidgetItem):
        old_text = item.text()
        new_text, ok = QInputDialog.getText(self, "Edit Role", "Role name:", QLineEdit.EchoMode.Normal, old_text)
        if ok and new_text.strip():
            item.setText(new_text.strip())
            self._save()
    
    def _on_edit_selected(self):
        item = self.list_widget.currentItem()
        if item:
            self._on_edit_item(item)
    
    def _on_remove_item(self):
        row = self.list_widget.currentRow()
        if row < 0:
            return
        item = self.list_widget.item(row)
        reply = QMessageBox.question(self, "Confirm", f"Remove '{item.text()}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
        if reply == QMessageBox.StandardButton.Yes:
            self.list_widget.takeItem(row)
            self._save()


class ProjectFieldsPanel(QFrame):
    """Panel for managing dynamic project field categories"""
    
    data_changed = pyqtSignal()
    
    def __init__(self):
        super().__init__()
        self._setup_ui()
        self.refresh()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(12)
        
        # Header
        header = QHBoxLayout()
        title_label = QLabel("Project Fields")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title_label.setFont(title_font)
        header.addWidget(title_label)
        
        header.addStretch()
        
        add_field_btn = QPushButton("+ Add Field")
        add_field_btn.clicked.connect(self._on_add_field)
        header.addWidget(add_field_btn)
        
        layout.addLayout(header)
        
        desc = QLabel("Define custom fields for your projects. Each field can have a list of allowed values.")
        desc.setStyleSheet("color: gray;")
        desc.setWordWrap(True)
        layout.addWidget(desc)
        
        # Scroll area for field cards
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        
        self.fields_container = QWidget()
        self.fields_layout = QVBoxLayout(self.fields_container)
        self.fields_layout.setContentsMargins(0, 0, 0, 0)
        self.fields_layout.setSpacing(12)
        self.fields_layout.addStretch()
        
        scroll.setWidget(self.fields_container)
        layout.addWidget(scroll, 1)
        
        hint = QLabel("Data: team_data.json -> project_fields[]")
        hint.setStyleSheet("color: gray; font-size: 10px;")
        layout.addWidget(hint)
    
    def refresh(self):
        """Refresh field cards from data"""
        # Clear existing cards (except the stretch at the end)
        while self.fields_layout.count() > 1:
            item = self.fields_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        if not data_service.team_data:
            return
        
        for i, field_data in enumerate(data_service.team_data.project_fields):
            if isinstance(field_data, dict):
                card = ProjectFieldCard(field_data, i)
                card.field_updated.connect(self._on_field_updated)
                card.field_deleted.connect(self._on_field_deleted)
                self.fields_layout.insertWidget(self.fields_layout.count() - 1, card)
    
    def _on_add_field(self):
        """Add a new project field"""
        dialog = AddFieldDialog(self)
        if dialog.exec() == QDialog.DialogCode.Accepted:
            key, label = dialog.get_data()
            
            # Check for duplicate key
            if data_service.team_data:
                for pf in data_service.team_data.project_fields:
                    if isinstance(pf, dict) and pf.get('key') == key:
                        QMessageBox.warning(self, "Duplicate", f"Field key '{key}' already exists.")
                        return
            
            new_field = {'key': key, 'label': label, 'values': []}
            data_service.team_data.project_fields.append(new_field)
            data_service.save_team_data(data_service.team_data)
            self.refresh()
            self.data_changed.emit()
    
    def _on_field_updated(self):
        """Handle field update"""
        self.data_changed.emit()
    
    def _on_field_deleted(self, index: int):
        """Handle field deletion"""
        if data_service.team_data and 0 <= index < len(data_service.team_data.project_fields):
            del data_service.team_data.project_fields[index]
            data_service.save_team_data(data_service.team_data)
            self.refresh()
            self.data_changed.emit()


class AddFieldDialog(QDialog):
    """Dialog for adding a new project field"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Add Project Field")
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QFormLayout(self)
        
        self.label_input = QLineEdit()
        self.label_input.setPlaceholderText("e.g., Department")
        self.label_input.textChanged.connect(self._update_key)
        layout.addRow("Label:", self.label_input)
        
        self.key_input = QLineEdit()
        self.key_input.setPlaceholderText("e.g., department")
        layout.addRow("Key:", self.key_input)
        
        hint = QLabel("Key is used internally. Auto-generated from label if left empty.")
        hint.setStyleSheet("color: gray; font-size: 10px;")
        layout.addRow("", hint)
        
        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        buttons.accepted.connect(self._validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addRow(buttons)
    
    def _update_key(self, text):
        """Auto-generate key from label"""
        if not self.key_input.text():
            key = text.lower().replace(' ', '_').replace('-', '_')
            key = ''.join(c for c in key if c.isalnum() or c == '_')
            self.key_input.setPlaceholderText(key or "e.g., department")
    
    def _validate_and_accept(self):
        label = self.label_input.text().strip()
        if not label:
            QMessageBox.warning(self, "Required", "Label is required.")
            return
        self.accept()
    
    def get_data(self) -> tuple:
        label = self.label_input.text().strip()
        key = self.key_input.text().strip()
        if not key:
            key = label.lower().replace(' ', '_').replace('-', '_')
            key = ''.join(c for c in key if c.isalnum() or c == '_')
        return key, label


class ProjectFieldCard(QFrame):
    """A card displaying a single project field with its values"""
    
    field_updated = pyqtSignal()
    field_deleted = pyqtSignal(int)
    
    def __init__(self, field_data: dict, index: int):
        super().__init__()
        self.field_data = field_data
        self.index = index
        
        self.setFrameStyle(QFrame.Shape.StyledPanel)
        self.setStyleSheet("""
            ProjectFieldCard {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 4px;
            }
        """)
        
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)
        
        # Header with label and actions
        header = QHBoxLayout()
        
        self.label_display = QLabel(self.field_data.get('label', 'Untitled'))
        label_font = QFont()
        label_font.setBold(True)
        self.label_display.setFont(label_font)
        header.addWidget(self.label_display)
        
        key_label = QLabel(f"({self.field_data.get('key', '')})")
        key_label.setStyleSheet("color: gray;")
        header.addWidget(key_label)
        
        header.addStretch()
        
        rename_btn = QPushButton("Rename")
        rename_btn.setFixedWidth(60)
        rename_btn.clicked.connect(self._on_rename)
        header.addWidget(rename_btn)
        
        delete_btn = QPushButton("Delete")
        delete_btn.setFixedWidth(60)
        delete_btn.clicked.connect(self._on_delete)
        header.addWidget(delete_btn)
        
        layout.addLayout(header)
        
        # Values list
        self.list_widget = QListWidget()
        self.list_widget.setMaximumHeight(150)
        self.list_widget.setAlternatingRowColors(True)
        self.list_widget.itemDoubleClicked.connect(self._on_edit_value)
        
        for val in self.field_data.get('values', []):
            self.list_widget.addItem(val)
        
        layout.addWidget(self.list_widget)
        
        # Add value row
        add_row = QHBoxLayout()
        
        self.new_value_input = QLineEdit()
        self.new_value_input.setPlaceholderText("Add value...")
        self.new_value_input.returnPressed.connect(self._on_add_value)
        add_row.addWidget(self.new_value_input)
        
        add_btn = QPushButton("+")
        add_btn.setFixedWidth(40)
        add_btn.clicked.connect(self._on_add_value)
        add_row.addWidget(add_btn)
        
        remove_btn = QPushButton("-")
        remove_btn.setFixedWidth(40)
        remove_btn.clicked.connect(self._on_remove_value)
        add_row.addWidget(remove_btn)
        
        import_btn = QPushButton("Import...")
        import_btn.clicked.connect(self._on_import)
        add_row.addWidget(import_btn)
        
        layout.addLayout(add_row)
    
    def _get_values(self) -> list:
        return [self.list_widget.item(i).text() for i in range(self.list_widget.count())]
    
    def _save(self):
        """Save values back to team_data"""
        if not data_service.team_data:
            return
        
        self.field_data['values'] = self._get_values()
        
        # Update in team_data
        if self.index < len(data_service.team_data.project_fields):
            data_service.team_data.project_fields[self.index] = self.field_data
            data_service.save_team_data(data_service.team_data)
            self.field_updated.emit()
    
    def _on_rename(self):
        """Rename the field label"""
        new_label, ok = QInputDialog.getText(
            self, "Rename Field", "Label:",
            QLineEdit.EchoMode.Normal,
            self.field_data.get('label', '')
        )
        
        if ok and new_label.strip():
            self.field_data['label'] = new_label.strip()
            self.label_display.setText(new_label.strip())
            self._save()
    
    def _on_delete(self):
        """Delete this field"""
        reply = QMessageBox.question(
            self,
            "Confirm Delete",
            f"Delete field '{self.field_data.get('label', '')}'?\n\nThis will remove the field and all its values.",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            self.field_deleted.emit(self.index)
    
    def _on_add_value(self):
        """Add a new value"""
        text = self.new_value_input.text().strip()
        if not text:
            return
        
        existing = self._get_values()
        if text in existing:
            QMessageBox.warning(self, "Duplicate", f"'{text}' already exists.")
            return
        
        self.list_widget.addItem(text)
        self.new_value_input.clear()
        self._save()
    
    def _on_edit_value(self, item: QListWidgetItem):
        """Edit a value"""
        old_text = item.text()
        new_text, ok = QInputDialog.getText(
            self, "Edit Value", "Value:",
            QLineEdit.EchoMode.Normal, old_text
        )
        
        if ok and new_text.strip():
            item.setText(new_text.strip())
            self._save()
    
    def _on_remove_value(self):
        """Remove selected value"""
        row = self.list_widget.currentRow()
        if row >= 0:
            self.list_widget.takeItem(row)
            self._save()
    
    def _on_import(self):
        """Import values from file"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Import Values",
            "",
            "All Supported (*.json *.csv *.xlsx *.xls);;JSON (*.json);;CSV (*.csv);;Excel (*.xlsx *.xls)"
        )
        
        if not file_path:
            return
        
        try:
            new_items = self._read_import_file(file_path)
            
            if not new_items:
                QMessageBox.information(self, "Import", "No items found.")
                return
            
            existing = set(self._get_values())
            added = 0
            
            for item in new_items:
                if item and item not in existing:
                    self.list_widget.addItem(item)
                    existing.add(item)
                    added += 1
            
            if added > 0:
                self._save()
            
            QMessageBox.information(self, "Import Complete", f"Added {added} values.")
            
        except Exception as e:
            QMessageBox.critical(self, "Import Error", str(e))
    
    def _read_import_file(self, file_path: str) -> list:
        """Read items from import file"""
        items = []
        
        if file_path.lower().endswith('.json'):
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            if isinstance(data, list):
                for item in data:
                    if isinstance(item, str):
                        items.append(item.strip())
                    elif isinstance(item, dict):
                        for key in ['name', 'value', 'label']:
                            if key in item:
                                items.append(str(item[key]).strip())
                                break
        
        elif file_path.lower().endswith('.csv'):
            with open(file_path, 'r', encoding='utf-8', newline='') as f:
                reader = csv.reader(f)
                for row in reader:
                    if row and row[0].strip():
                        items.append(row[0].strip())
        
        elif file_path.lower().endswith(('.xlsx', '.xls')):
            items = self._import_from_excel(file_path)
        
        return items
    
    def _import_from_excel(self, file_path: str) -> list:
        """Import from Excel with column picker"""
        try:
            import openpyxl
        except ImportError:
            QMessageBox.warning(self, "Missing Library", "Excel import requires openpyxl.")
            return []
        
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet = wb.active
        
        preview_rows = []
        for i, row in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
            preview_rows.append(row)
        
        if not preview_rows:
            return []
        
        dialog = ExcelColumnPickerDialog(preview_rows, self)
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return []
        
        col_index = dialog.selected_column
        skip_header = dialog.skip_header
        
        items = []
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if skip_header and i == 0:
                continue
            if col_index < len(row) and row[col_index]:
                items.append(str(row[col_index]).strip())
        
        wb.close()
        return items


class ThemePanel(QFrame):
    """Panel for selecting app theme"""
    
    theme_changed = pyqtSignal()
    
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self._setup_ui()
        self.refresh()
    
    def _setup_ui(self):
        from ..themes import theme_manager, THEMES
        
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(16)
        
        # Header
        title_label = QLabel("Theme")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title_label.setFont(title_font)
        layout.addWidget(title_label)
        
        desc = QLabel("Choose a color theme for the application. Changes apply immediately.")
        desc.setWordWrap(True)
        layout.addWidget(desc)
        
        # Theme buttons grid
        themes_frame = QFrame()
        themes_layout = QVBoxLayout(themes_frame)
        themes_layout.setSpacing(8)
        
        self.theme_buttons = {}
        
        theme_info = [
            ("blue", "Blue", "Blue buttons, orange accents, dark header"),
            ("red", "Red", "Red buttons, warm accents"),
            ("green", "Green", "Green buttons, teal accents"),
            ("mono", "Mono", "Black and white, minimal color"),
        ]
        
        for theme_id, name, description in theme_info:
            btn_frame = QFrame()
            btn_layout = QHBoxLayout(btn_frame)
            btn_layout.setContentsMargins(8, 8, 8, 8)
            
            # Color preview
            preview = QFrame()
            preview.setFixedSize(40, 40)
            theme = THEMES.get(theme_id)
            if theme:
                preview.setStyleSheet(f"""
                    QFrame {{
                        background-color: {theme.primary};
                        border: 2px solid {theme.header_bg};
                        border-radius: 4px;
                    }}
                """)
            btn_layout.addWidget(preview)
            
            # Theme info
            info_layout = QVBoxLayout()
            info_layout.setSpacing(2)
            
            name_label = QLabel(name)
            name_font = QFont()
            name_font.setBold(True)
            name_label.setFont(name_font)
            info_layout.addWidget(name_label)
            
            desc_label = QLabel(description)
            info_layout.addWidget(desc_label)
            
            btn_layout.addLayout(info_layout, 1)
            
            # Select button
            select_btn = QPushButton("Select")
            select_btn.setFixedWidth(80)
            select_btn.clicked.connect(lambda checked, tid=theme_id: self._on_select_theme(tid))
            btn_layout.addWidget(select_btn)
            
            self.theme_buttons[theme_id] = (btn_frame, select_btn)
            themes_layout.addWidget(btn_frame)
        
        layout.addWidget(themes_frame)
        
        # Current theme indicator
        self.current_label = QLabel()
        self.current_label.setStyleSheet("font-style: italic;")
        layout.addWidget(self.current_label)
        
        layout.addStretch()
        
        # Note
        note = QLabel("Theme preference is saved locally and won't affect other team members.")
        note.setWordWrap(True)
        layout.addWidget(note)
    
    def refresh(self):
        """Update the current theme indicator"""
        from ..themes import theme_manager
        
        current_id = theme_manager.current_theme_id
        current_name = theme_manager.current_theme.name
        self.current_label.setText(f"Current theme: {current_name}")
        
        # Update button states
        for theme_id, (frame, btn) in self.theme_buttons.items():
            if theme_id == current_id:
                btn.setText("Active")
                btn.setEnabled(False)
            else:
                btn.setText("Select")
                btn.setEnabled(True)
    
    def _on_select_theme(self, theme_id: str):
        """Handle theme selection"""
        from ..themes import theme_manager
        
        theme_manager.set_theme(theme_id)
        self.main_window.apply_theme()
        self.refresh()
        self.theme_changed.emit()


class OptionalTabPanel(QFrame):
    """Panel for configuring the optional 4th project tab"""
    
    config_changed = pyqtSignal()
    
    def __init__(self, main_window):
        super().__init__()
        self.main_window = main_window
        self._setup_ui()
        self.refresh()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(16)
        
        # Header
        title_label = QLabel("Optional Project Tab")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title_label.setFont(title_font)
        layout.addWidget(title_label)
        
        desc = QLabel(
            "Projects always have Details, Time Log, and Notes tabs. "
            "You can optionally enable a 4th tab for specialized content."
        )
        desc.setWordWrap(True)
        layout.addWidget(desc)
        
        # Enable checkbox
        self.enable_check = QCheckBox("Enable optional tab")
        self.enable_check.stateChanged.connect(self._on_enable_changed)
        layout.addWidget(self.enable_check)
        
        # Tab type selection
        type_frame = QFrame()
        type_layout = QFormLayout(type_frame)
        
        self.type_combo = QComboBox()
        self.type_combo.addItem("Chunking Guide", "chunking_guide")
        self.type_combo.currentIndexChanged.connect(self._on_type_changed)
        type_layout.addRow("Tab Type:", self.type_combo)
        
        self.label_edit = QLineEdit()
        self.label_edit.setPlaceholderText("Tab label...")
        self.label_edit.textChanged.connect(self._on_label_changed)
        type_layout.addRow("Tab Label:", self.label_edit)
        
        layout.addWidget(type_frame)
        
        # Description of tab type
        self.type_desc = QLabel()
        self.type_desc.setWordWrap(True)
        self.type_desc.setStyleSheet("color: gray; padding: 8px;")
        layout.addWidget(self.type_desc)
        
        layout.addStretch()
        
        # Note
        note = QLabel(
            "This setting is shared with your team. "
            "The optional tab will appear on all projects for all users."
        )
        note.setWordWrap(True)
        layout.addWidget(note)
    
    def refresh(self):
        """Load current settings"""
        if not data_service.team_data:
            self.enable_check.setChecked(False)
            self._update_type_enabled()
            return
        
        opt = data_service.team_data.optional_tab or {}
        
        self.enable_check.setChecked(opt.get('enabled', False))
        
        tab_type = opt.get('type', 'chunking_guide')
        idx = self.type_combo.findData(tab_type)
        if idx >= 0:
            self.type_combo.setCurrentIndex(idx)
        
        self.label_edit.setText(opt.get('label', 'Chunking Guide'))
        
        self._update_type_enabled()
        self._update_type_description()
    
    def _update_type_enabled(self):
        """Enable/disable type controls based on checkbox"""
        enabled = self.enable_check.isChecked()
        self.type_combo.setEnabled(enabled)
        self.label_edit.setEnabled(enabled)
    
    def _update_type_description(self):
        """Update description based on selected type"""
        tab_type = self.type_combo.currentData()
        
        descriptions = {
            'chunking_guide': (
                "Chunking Guide: Break courses into Training Modules (TMs) "
                "for production tracking. Each TM has a name and status. "
                "Ideal for L&D teams creating multi-module courses."
            ),
        }
        
        self.type_desc.setText(descriptions.get(tab_type, ''))
    
    def _save(self):
        """Save settings to team_data"""
        if not data_service.team_data:
            return
        
        data_service.team_data.optional_tab = {
            'enabled': self.enable_check.isChecked(),
            'type': self.type_combo.currentData(),
            'label': self.label_edit.text() or 'Chunking Guide',
        }
        
        data_service.save_team_data(data_service.team_data)
        
        # Notify that config changed
        self.config_changed.emit()
        
        # Rebuild tabs in project detail screen
        if hasattr(self.main_window, 'project_detail_screen'):
            self.main_window.project_detail_screen.rebuild_tabs()
    
    def _on_enable_changed(self):
        self._update_type_enabled()
        self._save()
    
    def _on_type_changed(self):
        self._update_type_description()
        self._save()
    
    def _on_label_changed(self):
        self._save()


class AdminScreen(QWidget):
    """Admin screen for managing team data and configuration"""
    
    def __init__(self, main_window):
        super().__init__()
        
        self.main_window = main_window
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Header
        header = self._create_header()
        layout.addWidget(header)
        
        # Main content with sidebar
        content = QSplitter(Qt.Orientation.Horizontal)
        
        # Panel stack (create BEFORE sidebar)
        self.panel_stack = QStackedWidget()
        self._create_panels()
        
        # Sidebar
        sidebar = self._create_sidebar()
        content.addWidget(sidebar)
        content.addWidget(self.panel_stack)
        content.setSizes([200, 600])
        
        layout.addWidget(content, 1)
        
        # Status bar
        status = self._create_status_bar()
        layout.addWidget(status)
    
    def _create_header(self):
        """Create the header"""
        header = QFrame()
        header.setObjectName("headerFrame")
        
        layout = QHBoxLayout(header)
        
        back_btn = QPushButton("<- Back")
        back_btn.clicked.connect(self._on_back)
        layout.addWidget(back_btn)
        
        title = QLabel("Admin Settings")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)
        
        subtitle = QLabel("Manage shared team data")
        layout.addWidget(subtitle)
        
        layout.addStretch()
        
        folder_btn = QPushButton("Data Folder...")
        folder_btn.clicked.connect(self._on_change_folder)
        layout.addWidget(folder_btn)
        
        self.folder_label = QLabel()
        self._update_folder_label()
        layout.addWidget(self.folder_label)
        
        return header
    
    def _create_sidebar(self):
        """Create the sidebar with category navigation"""
        sidebar = QFrame()
        sidebar.setObjectName("sidebarFrame")
        sidebar.setMaximumWidth(220)
        
        layout = QVBoxLayout(sidebar)
        layout.setContentsMargins(0, 8, 0, 8)
        layout.setSpacing(0)
        
        self.nav_list = QListWidget()
        
        # Navigation items
        categories = [
            ("-- TEAM --", None),
            ("  Employees", "employees"),
            ("  Team Roles", "team_roles"),
            ("-- WORK --", None),
            ("  Work Types", "work_types"),
            ("-- PROJECT --", None),
            ("  Custom Fields", "project_fields"),
            ("  Optional Tab", "optional_tab"),
            ("  Statuses", "project_statuses"),
            ("  Tags", "tags"),
            ("-- APP --", None),
            ("  Theme", "theme"),
        ]
        
        for label, key in categories:
            item = QListWidgetItem(label)
            if key is None:
                item.setFlags(Qt.ItemFlag.NoItemFlags)
                item.setForeground(Qt.GlobalColor.gray)
            else:
                item.setData(Qt.ItemDataRole.UserRole, key)
            self.nav_list.addItem(item)
        
        self.nav_list.currentItemChanged.connect(self._on_nav_changed)
        self.nav_list.setCurrentRow(1)
        
        layout.addWidget(self.nav_list, 1)
        
        return sidebar
    
    def _create_panels(self):
        """Create all the editing panels"""
        self.panels = {}
        
        # Employees panel
        self.panels['employees'] = EmployeesPanel()
        self.panel_stack.addWidget(self.panels['employees'])
        
        # Team roles panel
        self.panels['team_roles'] = TeamRolesPanel()
        self.panel_stack.addWidget(self.panels['team_roles'])
        
        # Work types panel
        self.panels['work_types'] = EditableListPanel('Work Types', 'work_types')
        self.panel_stack.addWidget(self.panels['work_types'])
        
        # Project fields panel (the big new one)
        self.panels['project_fields'] = ProjectFieldsPanel()
        self.panel_stack.addWidget(self.panels['project_fields'])
        
        # Optional tab panel
        self.panels['optional_tab'] = OptionalTabPanel(self.main_window)
        self.panel_stack.addWidget(self.panels['optional_tab'])
        
        # Project statuses panel
        self.panels['project_statuses'] = EditableListPanel('Project Statuses', 'project_statuses')
        self.panel_stack.addWidget(self.panels['project_statuses'])
        
        # Tags panel
        self.panels['tags'] = EditableListPanel('Tags', 'tags')
        self.panel_stack.addWidget(self.panels['tags'])
        
        # Theme panel
        self.panels['theme'] = ThemePanel(self.main_window)
        self.panel_stack.addWidget(self.panels['theme'])
    
    def _create_status_bar(self):
        """Create the status bar"""
        status = QFrame()
        status.setObjectName("footerFrame")
        
        layout = QHBoxLayout(status)
        layout.setContentsMargins(12, 4, 12, 4)
        
        self.status_label = QLabel("Editing: team_data.json")
        layout.addWidget(self.status_label)
        
        layout.addStretch()
        
        self.save_indicator = QLabel("Auto-saved")
        layout.addWidget(self.save_indicator)
        
        return status
    
    def _update_folder_label(self):
        """Update the data folder label"""
        if data_service.data_folder:
            path_str = str(data_service.data_folder)
            if len(path_str) > 40:
                path_str = "..." + path_str[-37:]
            self.folder_label.setText(path_str)
        else:
            self.folder_label.setText("No folder selected")
    
    def _on_nav_changed(self, current: QListWidgetItem, previous):
        """Handle navigation item change"""
        if not current:
            return
        
        key = current.data(Qt.ItemDataRole.UserRole)
        if key and key in self.panels:
            self.panel_stack.setCurrentWidget(self.panels[key])
    
    def _on_back(self):
        """Go back to home screen"""
        self.main_window.navigate_to("home")
    
    def _on_change_folder(self):
        """Change the data folder"""
        folder = QFileDialog.getExistingDirectory(
            self,
            "Select Data Folder",
            str(data_service.data_folder) if data_service.data_folder else "",
            QFileDialog.Option.ShowDirsOnly
        )
        
        if folder:
            data_service.data_folder = Path(folder)
            self._update_folder_label()
            self.refresh()
    
    def refresh(self):
        """Refresh all panels"""
        for panel in self.panels.values():
            panel.refresh()
        self._update_folder_label()
