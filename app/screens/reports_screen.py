"""
Reports Screen - View aggregated time data, export for external use
"""

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QFrame, QScrollArea, QComboBox, QDateEdit, QTableWidget,
    QTableWidgetItem, QHeaderView, QFileDialog, QMessageBox,
    QGroupBox, QProgressBar
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QFont
from datetime import datetime, date, timedelta
from pathlib import Path
import json
import csv

from ..data_service import data_service


class ReportsScreen(QWidget):
    """Reports screen with filters, summaries, and export"""
    
    def __init__(self, main_window):
        super().__init__()
        
        self.main_window = main_window
        self._entries = []  # Loaded time entries
        self._projects_cache = {}  # Project data cache
        
        self._setup_ui()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setSpacing(0)
        
        # Header
        header = self._create_header()
        layout.addWidget(header)
        
        # Filters bar
        filters = self._create_filters()
        layout.addWidget(filters)
        
        # Main content (scrollable)
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QFrame.Shape.NoFrame)
        
        content = QWidget()
        content_layout = QVBoxLayout(content)
        content_layout.setContentsMargins(20, 20, 20, 20)
        content_layout.setSpacing(20)
        
        # Summary cards
        summary = self._create_summary_section()
        content_layout.addWidget(summary)
        
        # By Project section
        by_project = self._create_by_project_section()
        content_layout.addWidget(by_project)
        
        # By Work Type section
        by_work_type = self._create_by_work_type_section()
        content_layout.addWidget(by_work_type)
        
        content_layout.addStretch()
        
        scroll.setWidget(content)
        layout.addWidget(scroll, 1)
        
        # Status bar
        status = self._create_status_bar()
        layout.addWidget(status)
    
    def _create_header(self):
        """Create header with back button and export buttons"""
        header = QFrame()
        header.setStyleSheet("""
            QFrame {
                background-color: #f8f9fa;
                border-bottom: 1px solid #dee2e6;
                padding: 12px;
            }
        """)
        
        layout = QHBoxLayout(header)
        
        # Back button
        back_btn = QPushButton("← Back")
        back_btn.clicked.connect(self._on_back)
        layout.addWidget(back_btn)
        
        # Title
        title = QLabel("Reports")
        title_font = QFont()
        title_font.setPointSize(14)
        title_font.setBold(True)
        title.setFont(title_font)
        layout.addWidget(title)
        
        layout.addStretch()
        
        # Export buttons
        csv_btn = QPushButton("📄 CSV")
        csv_btn.clicked.connect(lambda: self._on_export('csv'))
        layout.addWidget(csv_btn)
        
        excel_btn = QPushButton("📊 Excel")
        excel_btn.clicked.connect(lambda: self._on_export('excel'))
        layout.addWidget(excel_btn)
        
        json_btn = QPushButton("📋 JSON")
        json_btn.clicked.connect(lambda: self._on_export('json'))
        layout.addWidget(json_btn)
        
        return header
    
    def _create_filters(self):
        """Create the filters bar"""
        frame = QFrame()
        frame.setStyleSheet("""
            QFrame {
                background-color: white;
                border-bottom: 1px solid #dee2e6;
                padding: 12px 20px;
            }
        """)
        
        layout = QHBoxLayout(frame)
        
        # Period preset
        layout.addWidget(QLabel("Period:"))
        self.period_combo = QComboBox()
        self.period_combo.addItems([
            "Today",
            "Yesterday", 
            "This Week",
            "Last 7 Days",
            "This Month",
            "Last 30 Days",
            "This Quarter",
            "Custom"
        ])
        self.period_combo.setCurrentText("Last 7 Days")
        self.period_combo.currentTextChanged.connect(self._on_period_changed)
        layout.addWidget(self.period_combo)
        
        layout.addSpacing(20)
        
        # Date range
        layout.addWidget(QLabel("From:"))
        self.from_date = QDateEdit()
        self.from_date.setCalendarPopup(True)
        self.from_date.setDate(QDate.currentDate().addDays(-7))
        self.from_date.dateChanged.connect(self._on_custom_date)
        layout.addWidget(self.from_date)
        
        layout.addWidget(QLabel("To:"))
        self.to_date = QDateEdit()
        self.to_date.setCalendarPopup(True)
        self.to_date.setDate(QDate.currentDate())
        self.to_date.dateChanged.connect(self._on_custom_date)
        layout.addWidget(self.to_date)
        
        layout.addSpacing(20)
        
        # User filter
        layout.addWidget(QLabel("User:"))
        self.user_combo = QComboBox()
        self.user_combo.addItem("All Users", "all")
        self.user_combo.addItem("My Time", "me")
        self.user_combo.currentIndexChanged.connect(self._on_filter_changed)
        layout.addWidget(self.user_combo)
        
        layout.addStretch()
        
        # Apply button
        apply_btn = QPushButton("Apply")
        apply_btn.clicked.connect(self._load_data)
        layout.addWidget(apply_btn)
        
        return frame
    
    def _create_summary_section(self):
        """Create summary cards section"""
        frame = QFrame()
        layout = QHBoxLayout(frame)
        layout.setSpacing(20)
        
        self.total_hours_card = self._create_summary_card("0.0", "Total Hours")
        layout.addWidget(self.total_hours_card)
        
        self.projects_card = self._create_summary_card("0", "Projects")
        layout.addWidget(self.projects_card)
        
        self.avg_per_day_card = self._create_summary_card("0.0", "Avg/Day")
        layout.addWidget(self.avg_per_day_card)
        
        self.avg_ratio_card = self._create_summary_card("—", "Avg Ratio")
        layout.addWidget(self.avg_ratio_card)
        
        layout.addStretch()
        
        return frame
    
    def _create_summary_card(self, value: str, label: str) -> QFrame:
        """Create a single summary card"""
        card = QFrame()
        card.setStyleSheet("""
            QFrame {
                background-color: #f8f9fa;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                padding: 16px;
                min-width: 120px;
            }
        """)
        
        layout = QVBoxLayout(card)
        layout.setSpacing(4)
        layout.setContentsMargins(12, 12, 12, 12)
        
        value_label = QLabel(value)
        value_font = QFont()
        value_font.setPointSize(24)
        value_font.setBold(True)
        value_label.setFont(value_font)
        value_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(value_label)
        
        label_label = QLabel(label)
        label_label.setStyleSheet("color: #6c757d;")
        label_label.setAlignment(Qt.AlignmentFlag.AlignCenter)
        layout.addWidget(label_label)
        
        # Store reference for updates
        card.value_label = value_label
        
        return card
    
    def _create_by_project_section(self):
        """Create the By Project breakdown section"""
        group = QGroupBox("By Project")
        group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 8px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 8px;
            }
        """)
        
        layout = QVBoxLayout(group)
        
        self.project_table = QTableWidget()
        self.project_table.setColumnCount(4)
        self.project_table.setHorizontalHeaderLabels(["Project", "Hours", "Target", "Ratio"])
        self.project_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        self.project_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeMode.ResizeToContents)
        self.project_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        self.project_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        self.project_table.setAlternatingRowColors(True)
        self.project_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.project_table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.project_table.setMinimumHeight(200)
        
        layout.addWidget(self.project_table)
        
        return group
    
    def _create_by_work_type_section(self):
        """Create the By Work Type breakdown section"""
        group = QGroupBox("By Work Type")
        group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                font-size: 13px;
                border: 1px solid #dee2e6;
                border-radius: 8px;
                margin-top: 12px;
                padding-top: 8px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 12px;
                padding: 0 8px;
            }
        """)
        
        layout = QVBoxLayout(group)
        
        # Container for the bar chart
        self.work_type_container = QWidget()
        self.work_type_layout = QVBoxLayout(self.work_type_container)
        self.work_type_layout.setSpacing(8)
        
        layout.addWidget(self.work_type_container)
        
        return group
    
    def _create_status_bar(self):
        """Create status bar"""
        frame = QFrame()
        frame.setStyleSheet("""
            QFrame {
                background-color: #e9ecef;
                border-top: 1px solid #dee2e6;
                padding: 8px 16px;
            }
        """)
        
        layout = QHBoxLayout(frame)
        layout.setContentsMargins(0, 0, 0, 0)
        
        self.status_label = QLabel("No data loaded")
        self.status_label.setStyleSheet("color: #6c757d;")
        layout.addWidget(self.status_label)
        
        layout.addStretch()
        
        return frame
    
    def _on_back(self):
        """Go back to home"""
        self.main_window.navigate_to("home")
    
    def _on_period_changed(self, period: str):
        """Handle period preset change"""
        today = QDate.currentDate()
        
        if period == "Today":
            self.from_date.setDate(today)
            self.to_date.setDate(today)
        elif period == "Yesterday":
            yesterday = today.addDays(-1)
            self.from_date.setDate(yesterday)
            self.to_date.setDate(yesterday)
        elif period == "This Week":
            # Start of week (Monday)
            days_since_monday = today.dayOfWeek() - 1
            start = today.addDays(-days_since_monday)
            self.from_date.setDate(start)
            self.to_date.setDate(today)
        elif period == "Last 7 Days":
            self.from_date.setDate(today.addDays(-6))
            self.to_date.setDate(today)
        elif period == "This Month":
            start = QDate(today.year(), today.month(), 1)
            self.from_date.setDate(start)
            self.to_date.setDate(today)
        elif period == "Last 30 Days":
            self.from_date.setDate(today.addDays(-29))
            self.to_date.setDate(today)
        elif period == "This Quarter":
            quarter = (today.month() - 1) // 3
            start_month = quarter * 3 + 1
            start = QDate(today.year(), start_month, 1)
            self.from_date.setDate(start)
            self.to_date.setDate(today)
        # "Custom" - don't change dates
        
        if period != "Custom":
            self._load_data()
    
    def _on_custom_date(self):
        """Handle custom date change"""
        self.period_combo.setCurrentText("Custom")
    
    def _on_filter_changed(self):
        """Handle filter change"""
        self._load_data()
    
    def _load_data(self):
        """Load time entries based on current filters"""
        if not data_service.data_folder:
            return
        
        time_dir = data_service.data_folder / "time"
        if not time_dir.exists():
            self._entries = []
            self._update_display()
            return
        
        # Get date range
        from_date = self.from_date.date().toPyDate()
        to_date = self.to_date.date().toPyDate()
        
        # Get user filter
        user_filter = self.user_combo.currentData()
        current_user = data_service.current_user_id
        
        # Load entries
        self._entries = []
        self._projects_cache = {p.id: p for p in data_service.get_all_projects()}
        
        for file_path in time_dir.glob("*.json"):
            try:
                # Parse filename for user and date
                filename = file_path.stem  # e.g., "dsmith_2025-12-06"
                parts = filename.rsplit('_', 1)
                if len(parts) != 2:
                    continue
                
                file_user = parts[0]
                file_date_str = parts[1]
                
                try:
                    file_date = datetime.strptime(file_date_str, "%Y-%m-%d").date()
                except ValueError:
                    continue
                
                # Date filter
                if file_date < from_date or file_date > to_date:
                    continue
                
                # User filter
                if user_filter == "me" and file_user != current_user:
                    continue
                
                # Load file
                with open(file_path, 'r') as f:
                    data = json.load(f)
                
                for entry in data.get('entries', []):
                    entry['_user'] = file_user
                    entry['_date'] = file_date_str
                    self._entries.append(entry)
                    
            except (json.JSONDecodeError, IOError):
                continue
        
        self._update_display()
    
    def _update_display(self):
        """Update all display elements with current data"""
        self._update_summary()
        self._update_by_project()
        self._update_by_work_type()
        self._update_status()
    
    def _update_summary(self):
        """Update summary cards"""
        # Support both new 'hours' and legacy 'duration_minutes'
        total_hours = sum(
            e.get('hours', 0) if 'hours' in e else e.get('duration_minutes', 0) // 60
            for e in self._entries
        )
        
        project_ids = set(e.get('project_id') for e in self._entries)
        num_projects = len(project_ids)
        
        # Calculate days in range
        from_date = self.from_date.date().toPyDate()
        to_date = self.to_date.date().toPyDate()
        num_days = (to_date - from_date).days + 1
        avg_per_day = total_hours / num_days if num_days > 0 else 0
        
        # Calculate average ratio
        ratios = []
        for pid in project_ids:
            project = self._projects_cache.get(pid)
            if project and project.target_hours > 0:
                project_hours = sum(
                    e.get('hours', 0) if 'hours' in e else e.get('duration_minutes', 0) // 60
                    for e in self._entries 
                    if e.get('project_id') == pid
                )
                ratios.append(project_hours / project.target_hours)
        
        avg_ratio = sum(ratios) / len(ratios) if ratios else 0
        
        # Update cards
        self.total_hours_card.value_label.setText(f"{total_hours}")
        self.projects_card.value_label.setText(str(num_projects))
        self.avg_per_day_card.value_label.setText(f"{avg_per_day:.1f}")
        self.avg_ratio_card.value_label.setText(f"{avg_ratio:.2f}" if ratios else "—")
    
    def _update_by_project(self):
        """Update the By Project table"""
        # Group entries by project
        project_data = {}
        for entry in self._entries:
            pid = entry.get('project_id', 'Unknown')
            if pid not in project_data:
                project_data[pid] = {'hours': 0, 'entries': []}
            # Support both new 'hours' and legacy 'duration_minutes'
            hours = entry.get('hours', 0) if 'hours' in entry else entry.get('duration_minutes', 0) // 60
            project_data[pid]['hours'] += hours
            project_data[pid]['entries'].append(entry)
        
        # Sort by hours descending
        sorted_projects = sorted(
            project_data.items(),
            key=lambda x: x[1]['hours'],
            reverse=True
        )
        
        # Populate table
        self.project_table.setRowCount(len(sorted_projects))
        
        for row, (pid, data) in enumerate(sorted_projects):
            project = self._projects_cache.get(pid)
            name = project.name if project else pid
            hours = data['hours']
            target = project.target_hours if project else 0
            ratio = hours / target if target > 0 else 0
            
            self.project_table.setItem(row, 0, QTableWidgetItem(name))
            self.project_table.setItem(row, 1, QTableWidgetItem(f"{hours}"))
            self.project_table.setItem(row, 2, QTableWidgetItem(f"{target:.0f}" if target > 0 else "—"))
            
            ratio_item = QTableWidgetItem(f"{ratio:.2f}" if target > 0 else "—")
            if ratio > 1:
                ratio_item.setForeground(Qt.GlobalColor.red)
            self.project_table.setItem(row, 3, ratio_item)
    
    def _update_by_work_type(self):
        """Update the By Work Type bar chart"""
        # Clear existing bars
        while self.work_type_layout.count():
            item = self.work_type_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        # Group entries by work type
        work_type_data = {}
        for entry in self._entries:
            wt = entry.get('work_type', 'Unknown')
            if wt not in work_type_data:
                work_type_data[wt] = 0
            # Support both new 'hours' and legacy 'duration_minutes'
            hours = entry.get('hours', 0) if 'hours' in entry else entry.get('duration_minutes', 0) // 60
            work_type_data[wt] += hours
        
        if not work_type_data:
            label = QLabel("No data for selected period")
            label.setStyleSheet("color: #6c757d; padding: 20px;")
            label.setAlignment(Qt.AlignmentFlag.AlignCenter)
            self.work_type_layout.addWidget(label)
            return
        
        # Sort by hours descending
        sorted_types = sorted(work_type_data.items(), key=lambda x: x[1], reverse=True)
        total_hours = sum(work_type_data.values())
        max_hours = max(work_type_data.values())
        
        # Create bar for each work type
        for wt, hours in sorted_types:
            pct = (hours / total_hours * 100) if total_hours > 0 else 0
            bar_pct = (hours / max_hours * 100) if max_hours > 0 else 0
            
            row = QWidget()
            row_layout = QHBoxLayout(row)
            row_layout.setContentsMargins(0, 0, 0, 0)
            row_layout.setSpacing(12)
            
            # Label
            label = QLabel(wt)
            label.setMinimumWidth(150)
            label.setMaximumWidth(150)
            row_layout.addWidget(label)
            
            # Bar
            bar = QProgressBar()
            bar.setRange(0, 100)
            bar.setValue(int(bar_pct))
            bar.setTextVisible(False)
            bar.setStyleSheet("""
                QProgressBar {
                    border: none;
                    background-color: #e9ecef;
                    border-radius: 4px;
                    height: 20px;
                }
                QProgressBar::chunk {
                    background-color: #007bff;
                    border-radius: 4px;
                }
            """)
            row_layout.addWidget(bar, 1)
            
            # Stats
            stats = QLabel(f"{hours} hrs ({pct:.1f}%)")
            stats.setMinimumWidth(100)
            stats.setStyleSheet("color: #495057;")
            row_layout.addWidget(stats)
            
            self.work_type_layout.addWidget(row)
    
    def _update_status(self):
        """Update status bar"""
        from_str = self.from_date.date().toString("MMM d, yyyy")
        to_str = self.to_date.date().toString("MMM d, yyyy")
        
        project_count = len(set(e.get('project_id') for e in self._entries))
        
        self.status_label.setText(
            f"{len(self._entries)} entries across {project_count} projects  •  {from_str} – {to_str}"
        )
    
    def _on_export(self, format_type: str):
        """Export data to file"""
        if not self._entries:
            QMessageBox.information(self, "No Data", "No data to export.")
            return
        
        # Get save path
        if format_type == 'csv':
            filter_str = "CSV Files (*.csv)"
            default_ext = ".csv"
        elif format_type == 'excel':
            filter_str = "Excel Files (*.xlsx)"
            default_ext = ".xlsx"
        else:
            filter_str = "JSON Files (*.json)"
            default_ext = ".json"
        
        from_str = self.from_date.date().toString("yyyy-MM-dd")
        to_str = self.to_date.date().toString("yyyy-MM-dd")
        default_name = f"time_report_{from_str}_to_{to_str}{default_ext}"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Export Report",
            default_name,
            filter_str
        )
        
        if not file_path:
            return
        
        try:
            if format_type == 'csv':
                self._export_csv(file_path)
            elif format_type == 'excel':
                self._export_excel(file_path)
            else:
                self._export_json(file_path)
            
            QMessageBox.information(
                self,
                "Export Complete",
                f"Report exported to:\n{file_path}"
            )
        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Error",
                f"Failed to export:\n{str(e)}"
            )
    
    def _export_csv(self, file_path: str):
        """Export to CSV"""
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Date', 'User', 'Project ID', 'Project Name', 'Work Type', 'Hours', 'Notes'])
            
            for entry in sorted(self._entries, key=lambda e: e.get('_date', '')):
                pid = entry.get('project_id', '')
                project = self._projects_cache.get(pid)
                # Support both new 'hours' and legacy 'duration_minutes'
                hours = entry.get('hours', 0) if 'hours' in entry else entry.get('duration_minutes', 0) // 60
                
                writer.writerow([
                    entry.get('_date', ''),
                    entry.get('_user', ''),
                    pid,
                    project.name if project else '',
                    entry.get('work_type', ''),
                    hours,
                    entry.get('notes', '')
                ])
    
    def _export_excel(self, file_path: str):
        """Export to Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            # Fallback to CSV if openpyxl not installed
            csv_path = file_path.replace('.xlsx', '.csv')
            self._export_csv(csv_path)
            raise Exception(f"openpyxl not installed. Exported as CSV instead:\n{csv_path}")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Time Report"
        
        # Header
        headers = ['Date', 'User', 'Project ID', 'Project Name', 'Work Type', 'Hours', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Data
        for row, entry in enumerate(sorted(self._entries, key=lambda e: e.get('_date', '')), 2):
            pid = entry.get('project_id', '')
            project = self._projects_cache.get(pid)
            # Support both new 'hours' and legacy 'duration_minutes'
            hours = entry.get('hours', 0) if 'hours' in entry else entry.get('duration_minutes', 0) // 60
            
            ws.cell(row=row, column=1, value=entry.get('_date', ''))
            ws.cell(row=row, column=2, value=entry.get('_user', ''))
            ws.cell(row=row, column=3, value=pid)
            ws.cell(row=row, column=4, value=project.name if project else '')
            ws.cell(row=row, column=5, value=entry.get('work_type', ''))
            ws.cell(row=row, column=6, value=hours)
            ws.cell(row=row, column=7, value=entry.get('notes', ''))
        
        # Auto-size columns
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        wb.save(file_path)
    
    def _export_json(self, file_path: str):
        """Export to JSON (PowerBI-ready)"""
        export_data = []
        
        for entry in sorted(self._entries, key=lambda e: e.get('_date', '')):
            pid = entry.get('project_id', '')
            project = self._projects_cache.get(pid)
            # Support both new 'hours' and legacy 'duration_minutes'
            hours = entry.get('hours', 0) if 'hours' in entry else entry.get('duration_minutes', 0) // 60
            
            export_data.append({
                'date': entry.get('_date', ''),
                'user': entry.get('_user', ''),
                'project_id': pid,
                'project_name': project.name if project else '',
                'work_type': entry.get('work_type', ''),
                'hours': hours,
                'notes': entry.get('notes', '')
            })
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(export_data, f, indent=2)
    
    def refresh(self):
        """Refresh the screen"""
        # Populate user combo with employees
        current_data = self.user_combo.currentData()
        self.user_combo.clear()
        self.user_combo.addItem("All Users", "all")
        self.user_combo.addItem("My Time", "me")
        
        # Restore selection
        if current_data:
            idx = self.user_combo.findData(current_data)
            if idx >= 0:
                self.user_combo.setCurrentIndex(idx)
        
        self._load_data()
